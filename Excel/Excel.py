from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import os 

path = r'Excel\Data.xlsx'
if  os.path.exists(path):
    print("Path already exist !")
    workbook = load_workbook(path)
    worksheet = workbook.active
    table = worksheet.tables['Team_Agnie']
else:
    workbook = Workbook()
    workbook[workbook.sheetnames[0]].title = 'Boys'
    worksheet = workbook.active
    table = Table(displayName = 'Team_Agnie')
    table.tableStyleInfo = TableStyleInfo(name = 'TableStyleLight12', showFirstColumn=False, showLastColumn=False, showColumnStripes=True, showRowStripes=True )
    worksheet.add_table(table)
    table.ref = f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'

# worksheet['A1'] = 'Sri'
# worksheet['A2'] = 'Renga'
# worksheet['A3'] = 'Sarathy'
worksheet.append(['Intial','FirstName','LastName'])
worksheet.append(['R','Sri','Rengasarathy'])
worksheet.append(['R','Yuvaraj','Kannan'])
worksheet.append(['R','Naveen','Raja'])
worksheet.append(['P S','Kiruba','Nidhi'])
worksheet.append(['Y','Nikish','Daniel'])
worksheet.append(['J S','Prawin','Kumar'])
worksheet.append(['M','Sethu','Raj'])
table.ref = f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}'
print("Max Row : ",worksheet.max_column)
print("Max Column : ",worksheet.max_row)
print(f'A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}')

workbook.create_sheet('Girls')

# worksheet.delete_rows(1,3)
# worksheet.delete_rows(1)
# worksheet.delete_cols(1,3)
# worksheet.delete_cols(1)

print(workbook.sheetnames)

worksheet = workbook['Girls']
worksheet.append(['Intial','Name'])
worksheet.append(['G R','Harini'])
worksheet.append(['A','Prithika'])

worksheet = workbook['Boys']
# workbook.active = workbook.index(worksheet) # Activating the current sheet.
print(workbook.active)
workbook.create_sheet('Boys&Girls')
worksheet = workbook['Boys&Girls']
worksheet.append(['Intial','FirstName','LastName','Gender'])
genderValidator = DataValidation('list',formula1='Male,Female',allow_blank=True)
worksheet.add_data_validation(genderValidator)
genderValidator.add('D2:D1048576')
worksheet.append(['R','Sri','Rengasarathy'])
worksheet.append(['R','Yuvaraj','Kannan'])
worksheet.append(['R','Naveen','Raja'])
worksheet.append(['P S','Kiruba','Nidhi'])
worksheet.append(['Y','Nikish','Daniel'])
worksheet.append(['J S','Prawin','Kumar'])
worksheet.append(['M','Sethu','Raj'])
worksheet.append(['G R','Harini',''])
worksheet.append(['A','Prithika',''])
# workbook[workbook.sheetnames[0]].title = 'BOYS'
# worksheet.title = 'GIRLS'

# if 'Employee_Table' not in workbook.tables:
#     workbook.append(['Id','Name','Age','Languages Known'])
    
# else:
#     table = workbook.tables['Employee_Table']

for i in worksheet.iter_rows():
    for j in i:
        print(j.value,end=" ")
    print()
worksheet['A12'].comment = Comment(text="This is a Comment.", author="Sri")

workbook.save(path)
workbook.close()